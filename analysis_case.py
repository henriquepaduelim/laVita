from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
import json
import math

from openpyxl import load_workbook
import pandas as pd


BASE_DIR = Path(__file__).resolve().parent
REPORT_PATH = BASE_DIR / "analysis_case_report.md"
METRICS_PATH = BASE_DIR / "analysis_case_metrics.json"


@dataclass
class Agg:
    revenue: float = 0.0
    volume: float = 0.0
    stores: set | None = None
    products: set | None = None

    def __post_init__(self) -> None:
        if self.stores is None:
            self.stores = set()
        if self.products is None:
            self.products = set()

    def add(self, revenue: float, volume: float, store: str | None, product: str | None) -> None:
        self.revenue += revenue
        self.volume += volume
        if store:
            self.stores.add(store)
        if product:
            self.products.add(product)


def pct_change(current: float, previous: float) -> float | None:
    if previous == 0 or previous is None or pd.isna(previous):
        return None
    return (current - previous) / previous


def safe_ratio(numerator: float, denominator: float) -> float | None:
    if denominator == 0 or denominator is None or pd.isna(denominator):
        return None
    return numerator / denominator


def fmt_currency(value: float) -> str:
    return f"R$ {value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_number(value: float | int | None, digits: int = 0) -> str:
    if value is None or pd.isna(value):
        return "n/a"
    if digits == 0:
        return f"{int(round(float(value))):,}".replace(",", ".")
    return f"{float(value):,.{digits}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def fmt_pct(value: float | None, digits: int = 2) -> str:
    if value is None or pd.isna(value):
        return "n/a"
    return f"{value * 100:.{digits}f}%".replace(".", ",")


def normalize_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None or value == "":
        return None
    try:
        return pd.to_datetime(value).date()
    except Exception:
        return None


def load_dimensions() -> tuple[dict, dict, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    clients = pd.read_excel(BASE_DIR / "base_cliente.xlsx").rename(
        columns={
            "DS_REDE": "network",
            "NM_REDUZIDO": "store",
            "SG_ESTADO": "state",
            "NM_MUNICIPIO": "city",
        }
    )
    products = pd.read_excel(BASE_DIR / "base_produtos.xlsx").rename(
        columns={
            "CODIGO": "product_code",
            "PRODUTO": "product_name",
            "CATEGORIA": "category",
        }
    )
    macro = pd.read_excel(BASE_DIR / "base_macro_regiao.xlsx").rename(
        columns={"UF": "state", "MACRO_REGIAO": "macro_region"}
    )

    for col in ["network", "store", "state", "city"]:
        clients[col] = clients[col].astype(str).str.strip()
    for col in ["product_code", "product_name", "category"]:
        products[col] = products[col].astype(str).str.strip()
    for col in ["state", "macro_region"]:
        macro[col] = macro[col].astype(str).str.strip()

    client_map = clients.set_index("store")[["network", "state", "city"]].to_dict(orient="index")
    product_map = products.set_index("product_code")[["product_name", "category"]].to_dict(orient="index")
    return client_map, product_map, clients, products, macro


def aggregate_data(client_map: dict, product_map: dict) -> dict:
    workbook = load_workbook(BASE_DIR / "base_faturamento.xlsx", read_only=True, data_only=True)
    expected = ["DATA ENTREGA", "CLIENTE", "CD_PRODUTO2", "QUANTIDADE", "VENDA"]
    overall: dict[int, Agg] = defaultdict(Agg)
    network_year: dict[tuple[int, str], Agg] = defaultdict(Agg)
    category_year: dict[tuple[int, str], Agg] = defaultdict(Agg)
    quality = {
        "fact_rows": 0,
        "matched_store_rows": 0,
        "matched_product_rows": 0,
        "missing_stores": set(),
        "missing_products": set(),
        "min_date": {},
        "max_date": {},
        "year_store_sets": defaultdict(set),
        "year_network_sets": defaultdict(set),
        "year_product_sets": defaultdict(set),
    }

    for year_text in ("2024", "2025"):
        ws = workbook[year_text]
        headers = [str(v).strip() for v in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
        if headers != expected:
            workbook.close()
            raise ValueError(f"Cabecalho inesperado no faturamento[{year_text}]: {headers}")

        row_counter = 0
        for delivery_date, store_raw, product_raw, quantity_raw, revenue_raw in ws.iter_rows(
            min_row=2, values_only=True
        ):
            row_counter += 1
            year = int(year_text)
            store = str(store_raw).strip() if store_raw is not None else ""
            product_code = str(product_raw).strip() if product_raw is not None else ""
            quantity = float(quantity_raw or 0.0)
            revenue = float(revenue_raw or 0.0)
            d = normalize_date(delivery_date)

            client_info = client_map.get(store)
            product_info = product_map.get(product_code)
            network = client_info["network"] if client_info else None
            category = product_info["category"] if product_info else None

            quality["fact_rows"] += 1
            if client_info:
                quality["matched_store_rows"] += 1
                quality["year_store_sets"][year].add(store)
                quality["year_network_sets"][year].add(network)
            else:
                quality["missing_stores"].add(store)

            if product_info:
                quality["matched_product_rows"] += 1
                quality["year_product_sets"][year].add(product_code)
            else:
                quality["missing_products"].add(product_code)

            if d is not None:
                quality["min_date"][year] = min(d, quality["min_date"].get(year, d))
                quality["max_date"][year] = max(d, quality["max_date"].get(year, d))

            overall[year].add(revenue, quantity, store, product_code)
            if network:
                network_year[(year, network)].add(revenue, quantity, store, product_code)
            if category:
                category_year[(year, category)].add(revenue, quantity, store, product_code)

            if row_counter % 200000 == 0:
                print(f"[progress] {year_text}: {row_counter} linhas processadas", flush=True)

    workbook.close()

    return {
        "overall": overall,
        "network_year": network_year,
        "category_year": category_year,
        "quality": quality,
    }


def build_quality_snapshot(quality: dict, clients: pd.DataFrame, products: pd.DataFrame, macro: pd.DataFrame) -> dict:
    fact_rows = quality["fact_rows"]
    store_cov = (quality["matched_store_rows"] / fact_rows) * 100 if fact_rows else 0.0
    product_cov = (quality["matched_product_rows"] / fact_rows) * 100 if fact_rows else 0.0
    return {
        "fact_rows": fact_rows,
        "store_join_coverage_pct": round(store_cov, 4),
        "product_join_coverage_pct": round(product_cov, 4),
        "missing_store_count": len(quality["missing_stores"]),
        "missing_product_count": len(quality["missing_products"]),
        "missing_store_sample": sorted(quality["missing_stores"])[:15],
        "missing_product_sample": sorted(quality["missing_products"])[:15],
        "year_min_date": {str(k): str(v) for k, v in quality["min_date"].items()},
        "year_max_date": {str(k): str(v) for k, v in quality["max_date"].items()},
        "store_dim_unique_store": int(clients["store"].nunique()),
        "store_dim_unique_network": int(clients["network"].nunique()),
        "product_dim_unique_code": int(products["product_code"].nunique()),
        "macro_rows": int(len(macro)),
        "macro_unique_states": int(macro["state"].nunique()),
        "macro_unique_regions": int(macro["macro_region"].nunique()),
        "macro_duplicate_state_rows": int(macro["state"].duplicated().sum()),
        "client_states": sorted(clients["state"].dropna().unique().tolist()),
        "macro_states": sorted(macro["state"].dropna().unique().tolist()),
        "macro_join_status": (
            "Base de macro-regiao sem chave univoca suficiente para ligacao confiavel com loja ou municipio."
        ),
    }


def metric_bundle(agg: Agg, network_count: int | None = None) -> dict:
    avg_price = agg.revenue / agg.volume if agg.volume else 0.0
    return {
        "revenue": agg.revenue,
        "volume": agg.volume,
        "avg_price": avg_price,
        "assortment": len(agg.products),
        "active_stores": len(agg.stores),
        "active_networks": network_count,
    }


def build_network_table(network_year: dict[tuple[int, str], Agg]) -> pd.DataFrame:
    rows = []
    for (year, network), agg in network_year.items():
        rows.append(
            {
                "year": year,
                "network": network,
                "revenue": agg.revenue,
                "volume": agg.volume,
                "assortment": len(agg.products),
                "active_stores": len(agg.stores),
                "avg_unit_price": safe_ratio(agg.revenue, agg.volume),
                "ticket_per_store": safe_ratio(agg.revenue, len(agg.stores)),
            }
        )
    grouped = pd.DataFrame(rows)
    wide = grouped.pivot(index="network", columns="year")
    wide.columns = [f"{metric}_{year}" for metric, year in wide.columns]
    wide = wide.reset_index()
    needed = [
        "revenue_2024",
        "revenue_2025",
        "volume_2024",
        "volume_2025",
        "assortment_2024",
        "assortment_2025",
        "active_stores_2024",
        "active_stores_2025",
        "avg_unit_price_2024",
        "avg_unit_price_2025",
        "ticket_per_store_2024",
        "ticket_per_store_2025",
    ]
    for col in needed:
        if col not in wide.columns:
            wide[col] = 0.0
    wide["revenue_growth_abs"] = wide["revenue_2025"].fillna(0) - wide["revenue_2024"].fillna(0)
    wide["revenue_growth_pct"] = wide.apply(
        lambda row: pct_change(row["revenue_2025"], row["revenue_2024"]), axis=1
    )
    wide["is_new_2025"] = (wide["revenue_2024"].fillna(0) == 0) & (wide["revenue_2025"].fillna(0) > 0)
    return wide.sort_values("revenue_2025", ascending=False).reset_index(drop=True)


def build_category_table(category_year: dict[tuple[int, str], Agg]) -> pd.DataFrame:
    rows = []
    for (year, category), agg in category_year.items():
        rows.append(
            {
                "year": year,
                "category": category,
                "revenue": agg.revenue,
                "volume": agg.volume,
                "assortment": len(agg.products),
                "avg_unit_price": safe_ratio(agg.revenue, agg.volume),
            }
        )
    grouped = pd.DataFrame(rows)
    wide = grouped.pivot(index="category", columns="year")
    wide.columns = [f"{metric}_{year}" for metric, year in wide.columns]
    wide = wide.reset_index()
    needed = [
        "revenue_2024",
        "revenue_2025",
        "volume_2024",
        "volume_2025",
        "assortment_2024",
        "assortment_2025",
        "avg_unit_price_2024",
        "avg_unit_price_2025",
    ]
    for col in needed:
        if col not in wide.columns:
            wide[col] = 0.0
    wide["revenue_growth_abs"] = wide["revenue_2025"].fillna(0) - wide["revenue_2024"].fillna(0)
    wide["revenue_growth_pct"] = wide.apply(
        lambda row: pct_change(row["revenue_2025"], row["revenue_2024"]), axis=1
    )
    wide["volume_growth_pct"] = wide.apply(
        lambda row: pct_change(row["volume_2025"], row["volume_2024"]), axis=1
    )
    wide["price_growth_pct"] = wide.apply(
        lambda row: pct_change(row["avg_unit_price_2025"], row["avg_unit_price_2024"]), axis=1
    )
    return wide.sort_values("revenue_2025", ascending=False).reset_index(drop=True)


def build_pareto(network_table: pd.DataFrame) -> pd.DataFrame:
    df = network_table.sort_values("revenue_2025", ascending=False).copy()
    total = df["revenue_2025"].sum()
    df["revenue_share"] = df["revenue_2025"] / total
    df["cum_share"] = df["revenue_share"].cumsum()
    df["pareto_flag"] = False
    if not df.empty:
        cutoff = df["cum_share"].ge(0.8).idxmax()
        df.loc[:cutoff, "pareto_flag"] = True
    return df


def price_growth_relationship(network_table: pd.DataFrame) -> dict:
    comparable = network_table[
        (network_table["revenue_2024"].fillna(0) > 0)
        & (network_table["revenue_2025"].fillna(0) > 0)
        & network_table["avg_unit_price_2025"].notna()
        & network_table["revenue_growth_pct"].notna()
    ].copy()
    if len(comparable) < 2:
        return {
            "correlation_pearson": None,
            "top_growth_network": None,
            "top_growth_has_max_price": None,
            "max_price_network": None,
        }

    corr = comparable["revenue_growth_pct"].corr(comparable["avg_unit_price_2025"])
    top_growth = comparable.sort_values("revenue_growth_pct", ascending=False).iloc[0]
    max_price = comparable.sort_values("avg_unit_price_2025", ascending=False).iloc[0]
    return {
        "correlation_pearson": None if pd.isna(corr) else float(corr),
        "top_growth_network": str(top_growth["network"]),
        "top_growth_pct": float(top_growth["revenue_growth_pct"]),
        "top_growth_price": float(top_growth["avg_unit_price_2025"]),
        "max_price_network": str(max_price["network"]),
        "max_price": float(max_price["avg_unit_price_2025"]),
        "top_growth_has_max_price": bool(top_growth["network"] == max_price["network"]),
    }


def build_report(
    quality_snapshot: dict,
    overall: dict[int, dict],
    network_table: pd.DataFrame,
    pareto: pd.DataFrame,
    category_table: pd.DataFrame,
    price_growth: dict,
) -> str:
    y24 = overall[2024]
    y25 = overall[2025]
    pareto_clients = pareto.loc[pareto["pareto_flag"]]
    comparable = network_table[network_table["revenue_2024"].fillna(0) > 0].copy()
    drivers = comparable.sort_values("revenue_growth_abs", ascending=False).head(5)
    detractors = comparable.sort_values("revenue_growth_abs", ascending=True).head(5)
    top_ticket = network_table.sort_values("ticket_per_store_2025", ascending=False).iloc[0]
    top_volume = network_table.sort_values("volume_2025", ascending=False).iloc[0]
    top_assortment = network_table.sort_values("assortment_2025", ascending=False).iloc[0]
    top_growth = comparable.sort_values("revenue_growth_pct", ascending=False).iloc[0]
    top_categories = category_table.head(6)

    lines: list[str] = []
    lines.append("# Analise Do Case La Vita")
    lines.append("")
    lines.append("## 1. Decisoes Metodologicas")
    lines.append("- Cliente foi tratado como `rede`; loja foi tratada como `NM_REDUZIDO`.")
    lines.append("- Pareto e comparativos principais usam `2025` como ano corrente; `2024` entra como base para YoY.")
    lines.append("- Ticket medio por loja = faturamento da rede em 2025 / numero de lojas ativas da rede em 2025.")
    lines.append("- Preco medio = faturamento total / quantidade total; nao foi usada media simples de preco de linha.")
    lines.append("- Crescimento percentual por rede considera apenas redes com base comparavel (`receita 2024 > 0`).")
    lines.append("- Macro-regiao nao foi usada como dimensao principal porque a base nao permite ligacao deterministica por loja ou municipio.")
    lines.append("")
    lines.append("## 2. Qualidade Dos Dados")
    lines.append(f"- Linhas consolidadas no faturamento: `{fmt_number(quality_snapshot['fact_rows'])}`.")
    lines.append(f"- Cobertura do join de lojas: `{fmt_number(quality_snapshot['store_join_coverage_pct'], 4)}%`.")
    lines.append(f"- Cobertura do join de produtos: `{fmt_number(quality_snapshot['product_join_coverage_pct'], 4)}%`.")
    lines.append(
        f"- Faixa temporal 2024: `{quality_snapshot['year_min_date']['2024']}` a `{quality_snapshot['year_max_date']['2024']}`; "
        f"2025: `{quality_snapshot['year_min_date']['2025']}` a `{quality_snapshot['year_max_date']['2025']}`."
    )
    lines.append(
        f"- Cadastro de clientes: `{fmt_number(quality_snapshot['store_dim_unique_store'])}` lojas em `{fmt_number(quality_snapshot['store_dim_unique_network'])}` redes."
    )
    lines.append(
        f"- Cadastro de produtos: `{fmt_number(quality_snapshot['product_dim_unique_code'])}` codigos unicos."
    )
    lines.append("- A base de macro-regiao foi considerada inadequada para analise principal por falta de chave de ligacao confiavel.")
    lines.append("")
    lines.append("## 3. Visao Geral YoY")
    lines.append(
        f"- Receita 2024: `{fmt_currency(y24['revenue'])}`; 2025: `{fmt_currency(y25['revenue'])}`; variacao `{fmt_pct(pct_change(y25['revenue'], y24['revenue']))}`."
    )
    lines.append(
        f"- Volume 2024: `{fmt_number(y24['volume'])}`; 2025: `{fmt_number(y25['volume'])}`; variacao `{fmt_pct(pct_change(y25['volume'], y24['volume']))}`."
    )
    lines.append(
        f"- Preco medio 2024: `{fmt_currency(y24['avg_price'])}`; 2025: `{fmt_currency(y25['avg_price'])}`; variacao `{fmt_pct(pct_change(y25['avg_price'], y24['avg_price']))}`."
    )
    lines.append(
        f"- Sortimento 2024: `{fmt_number(y24['assortment'])}`; 2025: `{fmt_number(y25['assortment'])}`; variacao `{fmt_pct(pct_change(y25['assortment'], y24['assortment']))}`."
    )
    lines.append(
        f"- Lojas ativas 2024: `{fmt_number(y24['active_stores'])}`; 2025: `{fmt_number(y25['active_stores'])}`."
    )
    lines.append("")
    lines.append("## 4. Curva ABC / Pareto")
    lines.append(f"- `{fmt_number(len(pareto_clients))}` redes concentram `80%` do faturamento de 2025.")
    for _, row in pareto_clients.head(15).iterrows():
        lines.append(
            f"- `{row['network']}`: receita `{fmt_currency(row['revenue_2025'])}` | acumulado `{fmt_pct(row['cum_share'])}`."
        )
    lines.append("")
    lines.append("## 5. Analise Comparativa Por Cliente")
    lines.append(
        f"- Maior ticket medio por loja: `{top_ticket['network']}` com `{fmt_currency(top_ticket['ticket_per_store_2025'])}`."
    )
    lines.append(
        f"- Maior volume de itens em 2025: `{top_volume['network']}` com `{fmt_number(top_volume['volume_2025'])}` unidades."
    )
    lines.append(
        f"- Maior sortimento em 2025: `{top_assortment['network']}` com `{fmt_number(top_assortment['assortment_2025'])}` SKUs."
    )
    lines.append(
        f"- Maior crescimento comparavel em 2025: `{top_growth['network']}` com `{fmt_pct(top_growth['revenue_growth_pct'])}`."
    )
    lines.append(
        f"- Rede com maior preco medio em 2025: `{price_growth['max_price_network']}` com `{fmt_currency(price_growth['max_price'])}`."
    )
    lines.append(
        f"- A rede de maior crescimento foi `{price_growth['top_growth_network']}` com preco medio de `{fmt_currency(price_growth['top_growth_price'])}`. "
        f"Mesmo cliente com maior preco medio: `{str(price_growth['top_growth_has_max_price']).lower()}`."
    )
    lines.append(
        f"- Correlacao entre crescimento e preco medio em redes comparaveis: `{fmt_number(price_growth['correlation_pearson'], 4)}`."
    )
    lines.append("")
    lines.append("## 6. Dimensao Escolhida: Categoria")
    lines.append("- Categoria foi escolhida por ter chave completa produto->categoria e menor risco de ambiguidade.")
    for _, row in top_categories.iterrows():
        lines.append(
            f"- `{row['category']}`: receita 2025 `{fmt_currency(row['revenue_2025'])}` | crescimento `{fmt_pct(row['revenue_growth_pct'])}` | "
            f"volume `{fmt_pct(row['volume_growth_pct'])}` | preco `{fmt_pct(row['price_growth_pct'])}`."
        )
    lines.append("")
    lines.append("## 7. YoY Em 3 Blocos")
    lines.append("### Bloco 1 - Visao Geral")
    lines.append("- Receita, volume e preco medio indicam se o crescimento veio de expansao real ou apenas de repasse/preco.")
    lines.append("### Bloco 2 - Impulsionadores E Detratores")
    lines.append("- Top 5 impulsionadores por rede:")
    for _, row in drivers.iterrows():
        lines.append(
            f"- `{row['network']}`: delta `{fmt_currency(row['revenue_growth_abs'])}` | crescimento `{fmt_pct(row['revenue_growth_pct'])}`."
        )
    lines.append("- Top 5 detratores por rede:")
    for _, row in detractors.iterrows():
        lines.append(
            f"- `{row['network']}`: delta `{fmt_currency(row['revenue_growth_abs'])}` | crescimento `{fmt_pct(row['revenue_growth_pct'])}`."
        )
    lines.append("### Bloco 3 - Recomendacoes")
    lines.append("- Replicar alavancas das redes comparaveis que cresceram com preco sustentado.")
    lines.append("- Expandir sortimento nas redes medias com boa base de volume e baixa diversidade de SKU.")
    lines.append("- Separar clientes novos da analise de crescimento organico.")
    lines.append("- Monitorar erosao de preco em contas que crescem abaixo do mercado ou perdem receita.")
    lines.append("")
    lines.append("## 8. Modelo De Remuneracao Variavel")
    lines.append("- Peso 40%: crescimento de volume comparavel YoY.")
    lines.append("- Peso 30%: disciplina de preco, medindo preco medio versus meta ou faixa de referencia.")
    lines.append("- Peso 20%: ganho de sortimento por cliente.")
    lines.append("- Peso 10%: execucao de mix em clientes foco.")
    lines.append("- Regras de protecao: piso de faturamento, trilha separada para clientes novos e redutor para crescimento baseado apenas em desconto.")
    lines.append("")
    lines.append("## 9. Ferramentas, Frameworks E Metodos")
    lines.append("- `Python` para ingestao e reproducao.")
    lines.append("- `pandas` para consolidacao final e tabelas analiticas.")
    lines.append("- `openpyxl` em leitura streaming para planilhas grandes.")
    lines.append("- `DuckDB` recomendado para versao futura com auditoria SQL e performance melhor em agregacoes.")
    lines.append("- `pandera` ou validacoes explicitas para esquema, ranges e unicidade.")
    lines.append("- Metodos: profiling de dados, validacao de joins, reconciliacao de totais, definicao formal de metricas, Pareto, YoY comparavel e analise por dimensao com chave confiavel.")
    lines.append("")
    return "\n".join(lines) + "\n"


def main() -> None:
    client_map, product_map, clients, products, macro = load_dimensions()
    aggregate = aggregate_data(client_map, product_map)
    quality_snapshot = build_quality_snapshot(aggregate["quality"], clients, products, macro)

    overall = {
        2024: metric_bundle(aggregate["overall"][2024], len(aggregate["quality"]["year_network_sets"][2024])),
        2025: metric_bundle(aggregate["overall"][2025], len(aggregate["quality"]["year_network_sets"][2025])),
    }
    network_table = build_network_table(aggregate["network_year"])
    category_table = build_category_table(aggregate["category_year"])
    pareto = build_pareto(network_table)
    price_growth = price_growth_relationship(network_table)
    report = build_report(quality_snapshot, overall, network_table, pareto, category_table, price_growth)

    metrics = {
        "quality": quality_snapshot,
        "overall": overall,
        "pareto_80_networks": pareto.loc[pareto["pareto_flag"], ["network", "revenue_2025", "cum_share"]].to_dict(
            orient="records"
        ),
        "top_ticket_per_store_2025": network_table.sort_values("ticket_per_store_2025", ascending=False)
        .head(10)
        .to_dict(orient="records"),
        "top_volume_2025": network_table.sort_values("volume_2025", ascending=False).head(10).to_dict(orient="records"),
        "top_assortment_2025": network_table.sort_values("assortment_2025", ascending=False)
        .head(10)
        .to_dict(orient="records"),
        "top_growth_comparable_2025": network_table[network_table["revenue_2024"].fillna(0) > 0]
        .sort_values("revenue_growth_pct", ascending=False)
        .head(10)
        .to_dict(orient="records"),
        "category_summary": category_table.head(15).to_dict(orient="records"),
        "price_growth_relationship": price_growth,
    }

    REPORT_PATH.write_text(report, encoding="utf-8")
    METRICS_PATH.write_text(json.dumps(metrics, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
    print(report)
    print(f"Relatorio salvo em: {REPORT_PATH}")
    print(f"Metricas salvas em: {METRICS_PATH}")


if __name__ == "__main__":
    main()
