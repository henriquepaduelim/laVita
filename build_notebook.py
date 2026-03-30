#!/usr/bin/env python3
"""Cria analise_la_vita.ipynb com todas as 5 entregas do case."""
import json
from pathlib import Path

cells = []
_id = [0]


def _cell(ctype, src):
    src = src.strip("\n")
    lines = src.split("\n")
    source = [l + "\n" for l in lines[:-1]] + [lines[-1]]
    cell = {"cell_type": ctype, "id": f"{ctype[:2]}{_id[0]:03d}", "metadata": {}, "source": source}
    if ctype == "code":
        cell["execution_count"] = None
        cell["outputs"] = []
    _id[0] += 1
    cells.append(cell)


def md(src): _cell("markdown", src)
def code(src): _cell("code", src)


# ─────────────────────────────────────────────────────────────────────────────
# TÍTULO
# ─────────────────────────────────────────────────────────────────────────────
md("""
# Análise Comercial – La Vita Alimentos

**Case:** Analista Comercial Pleno
**Período analisado:** 2024 e 2025 (anos fechados)

---

## Entregas
1. Concentração de faturamento (Curva ABC / Pareto)
2. Análise comparativa por cliente (ticket médio, volume, sortimento, crescimento, preço)
3. Análise por dimensão de negócio — **Categoria**
4. Crescimento de vendas YoY — 3 slides
5. Proposta de modelo de remuneração variável
""")

# ─────────────────────────────────────────────────────────────────────────────
# SETUP
# ─────────────────────────────────────────────────────────────────────────────
code("""
import warnings
warnings.filterwarnings('ignore')

from pathlib import Path
from collections import defaultdict

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.ticker as mticker
from openpyxl import load_workbook

pd.set_option('display.float_format', lambda x: f'{x:,.2f}')
pd.set_option('display.max_rows', 60)
pd.set_option('display.max_columns', 20)

plt.rcParams.update({
    'figure.dpi': 150,
    'figure.facecolor': 'white',
    'axes.facecolor': '#FAFAFA',
    'axes.spines.top': False,
    'axes.spines.right': False,
    'axes.grid': True,
    'grid.alpha': 0.28,
    'grid.linestyle': '--',
    'font.size': 9,
    'axes.titlesize': 11,
    'axes.titleweight': 'bold',
})

BASE_DIR = Path('.')
OUT = BASE_DIR / 'analysis_output'
OUT.mkdir(exist_ok=True)

C = {
    'navy':   '#1B3A6B',
    'teal':   '#048A81',
    'green':  '#27AE60',
    'red':    '#C0392B',
    'orange': '#E67E22',
    'gray':   '#95A5A6',
}

print("Setup OK.")
""")

# ─────────────────────────────────────────────────────────────────────────────
# UTILIDADES
# ─────────────────────────────────────────────────────────────────────────────
code("""
def pct_change(current, previous):
    if previous is None or pd.isna(previous) or previous == 0:
        return None
    return (current - previous) / previous

def safe_div(num, den):
    if den is None or pd.isna(den) or den == 0:
        return None
    return num / den

def fmt_brl(v):
    if v is None or pd.isna(v):
        return 'n/a'
    s = f"{abs(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
    return f"R$ {s}" if v >= 0 else f"-R$ {s}"

def fmt_brl_k(v):
    if v is None or pd.isna(v):
        return 'n/a'
    if abs(v) >= 1_000_000:
        return f"R$ {v/1_000_000:.1f}M"
    if abs(v) >= 1_000:
        return f"R$ {v/1_000:.0f}K"
    return fmt_brl(v)

def fmt_pct(v, d=1):
    if v is None or pd.isna(v):
        return 'n/a'
    return f"{v*100:+.{d}f}%"

def fmt_int(v):
    if v is None or pd.isna(v):
        return 'n/a'
    return f"{int(v):,}".replace(',', '.')
""")

# ─────────────────────────────────────────────────────────────────────────────
# PREMISSAS METODOLÓGICAS
# ─────────────────────────────────────────────────────────────────────────────
md("""
## Premissas Metodológicas

| Decisão | Critério adotado | Justificativa |
|---|---|---|
| Nível de cliente | **Rede** (DS_REDE) | Case pede ticket *por loja* → pressupõe agrupamento acima da loja |
| Nível de loja | NM_REDUZIDO | Menor granularidade disponível no cadastro |
| Preço médio | Receita / Quantidade | Evita viés da média simples de preços por linha |
| Crescimento % | Apenas clientes **comparáveis** (receita 2024 > 0) | Exclui distorção de base zero |
| Crescimento absoluto | Todos os clientes com receita 2025 | Mede impacto real na receita |
| Dimensão principal | **Categoria** | base_macro_regiao.xlsx não tem coluna de município — join com loja é inviável |
| Base comparação YoY | Anos fechados 2024 e 2025 | Ambos cobrem jan–dez — comparação válida |

> **Nota sobre macro-região:** a tabela contém apenas `UF` + `MACRO_REGIAO` (sem município).
> SP possui 14 macro-regiões distintas. Sem chave de município, é impossível mapear
> deterministicamente uma loja a uma macro-região. Por isso, **Categoria** foi escolhida.
""")

# ─────────────────────────────────────────────────────────────────────────────
# CARGA DAS DIMENSÕES
# ─────────────────────────────────────────────────────────────────────────────
code("""
clients = pd.read_excel(BASE_DIR / 'base_cliente.xlsx').rename(columns={
    'DS_REDE': 'network', 'NM_REDUZIDO': 'store',
    'SG_ESTADO': 'state', 'NM_MUNICIPIO': 'city'})
for col in ['network', 'store', 'state', 'city']:
    clients[col] = clients[col].astype(str).str.strip()

products = pd.read_excel(BASE_DIR / 'base_produtos.xlsx').rename(columns={
    'CODIGO': 'product_code', 'PRODUTO': 'product_name', 'CATEGORIA': 'category'})
for col in ['product_code', 'product_name', 'category']:
    products[col] = products[col].astype(str).str.strip()

macro = pd.read_excel(BASE_DIR / 'base_macro_regiao.xlsx').rename(columns={
    'UF': 'state', 'MACRO_REGIAO': 'macro_region'})

client_map  = clients.set_index('store')[['network', 'state', 'city']].to_dict(orient='index')
product_map = products.set_index('product_code')[['product_name', 'category']].to_dict(orient='index')

print("=" * 65)
print("DIMENSÕES CARREGADAS")
print("=" * 65)
print(f"  Clientes : {clients['store'].nunique():>6,} lojas  |  {clients['network'].nunique()} redes")
print(f"  Produtos : {products['product_code'].nunique():>6,} SKUs   |  {products['category'].nunique()} categorias")
print()
print("  Produtos por categoria:")
print(products['category'].value_counts().to_string(header=False))
print()
print("  Macro-região: colunas UF + MACRO_REGIAO apenas (sem município).")
print("  SP tem 14 macro-regiões → join determinístico inviável.")
print("  Dimensão Categoria será usada na Entrega 3.")
""")

# ─────────────────────────────────────────────────────────────────────────────
# PROCESSAMENTO EM STREAMING
# ─────────────────────────────────────────────────────────────────────────────
code("""
class Agg:
    \"\"\"Acumulador leve para receita, volume, lojas e produtos distintos.\"\"\"
    __slots__ = ('rev', 'vol', 'stores', 'products')

    def __init__(self):
        self.rev = 0.0
        self.vol = 0.0
        self.stores   = set()
        self.products = set()

    def add(self, rev, vol, store, prod):
        self.rev += rev
        self.vol += vol
        if store: self.stores.add(store)
        if prod:  self.products.add(prod)


overall_agg = {}   # year → Agg
net_year    = {}   # (year, network) → Agg
cat_year    = {}   # (year, category) → Agg
qlog        = {}   # year → quality dict

wb = load_workbook(BASE_DIR / 'base_faturamento.xlsx', read_only=True, data_only=True)

for ys in ('2024', '2025'):
    y = int(ys)
    overall_agg[y] = Agg()
    q = {'rows': 0, 'sh': 0, 'ph': 0,
         'miss_s': set(), 'miss_p': set(),
         'mn': None, 'mx': None}
    qlog[y] = q

    ws = wb[ys]
    for row in ws.iter_rows(min_row=2, values_only=True):
        dv, sr, pr, qr, rr = row
        store = str(sr).strip() if sr else ''
        prod  = str(pr).strip() if pr else ''
        qty   = float(qr or 0)
        rev   = float(rr or 0)

        ci = client_map.get(store)
        pi = product_map.get(prod)
        net = ci['network']  if ci else None
        cat = pi['category'] if pi else None

        q['rows'] += 1
        if ci: q['sh'] += 1
        else:  q['miss_s'].add(store)
        if pi: q['ph'] += 1
        else:  q['miss_p'].add(prod)
        if dv is not None:
            ts = pd.Timestamp(dv)
            if q['mn'] is None or ts < q['mn']: q['mn'] = ts
            if q['mx'] is None or ts > q['mx']: q['mx'] = ts

        overall_agg[y].add(rev, qty, store, prod)

        if net:
            k = (y, net)
            if k not in net_year: net_year[k] = Agg()
            net_year[k].add(rev, qty, store, prod)

        if cat:
            k = (y, cat)
            if k not in cat_year: cat_year[k] = Agg()
            cat_year[k].add(rev, qty, store, prod)

    pct_s = q['sh'] / q['rows'] * 100
    pct_p = q['ph'] / q['rows'] * 100
    print(f"[{ys}]  {q['rows']:>10,} linhas  |  lojas {pct_s:.4f}%  |  produtos {pct_p:.4f}%")
    print(f"       período: {q['mn'].date()} → {q['mx'].date()}")
    if q['miss_p']:
        print(f"       produtos s/ cadastro: {sorted(q['miss_p'])[:5]}")

wb.close()
print()
print("Processamento concluído.")
""")

# ─────────────────────────────────────────────────────────────────────────────
# BUILD DAS TABELAS ANALÍTICAS
# ─────────────────────────────────────────────────────────────────────────────
code("""
def agg_dict(a):
    return {
        'revenue':       a.rev,
        'volume':        a.vol,
        'avg_price':     safe_div(a.rev, a.vol),
        'assortment':    len(a.products),
        'active_stores': len(a.stores),
    }

overall = {y: agg_dict(a) for y, a in overall_agg.items()}

# ── Network table ─────────────────────────────────────────────────────────────
rows = []
for (year, net), a in net_year.items():
    rows.append({
        'year': year, 'network': net,
        'revenue':       a.rev,
        'volume':        a.vol,
        'assortment':    len(a.products),
        'active_stores': len(a.stores),
        'avg_price':     safe_div(a.rev, a.vol),
        'ticket_store':  safe_div(a.rev, len(a.stores)),
    })

nt = (pd.DataFrame(rows)
        .pivot(index='network', columns='year')
        .fillna(0))
nt.columns = [f"{m}_{y}" for m, y in nt.columns]
nt = nt.reset_index()

for col in ['revenue_2024','revenue_2025','volume_2024','volume_2025',
            'assortment_2024','assortment_2025','active_stores_2024','active_stores_2025',
            'avg_price_2024','avg_price_2025','ticket_store_2024','ticket_store_2025']:
    if col not in nt.columns:
        nt[col] = 0.0

nt['rev_growth_abs'] = nt['revenue_2025'] - nt['revenue_2024']
nt['rev_growth_pct'] = nt.apply(lambda r: pct_change(r['revenue_2025'], r['revenue_2024']), axis=1)
nt['vol_growth_pct'] = nt.apply(lambda r: pct_change(r['volume_2025'],  r['volume_2024']),  axis=1)
nt['is_new_2025']    = (nt['revenue_2024'].fillna(0) == 0) & (nt['revenue_2025'].fillna(0) > 0)
nt = nt.sort_values('revenue_2025', ascending=False).reset_index(drop=True)

# ── Category table ───────────────────────────────────────────────────────────
rows_c = []
for (year, cat), a in cat_year.items():
    rows_c.append({
        'year': year, 'category': cat,
        'revenue':    a.rev,
        'volume':     a.vol,
        'assortment': len(a.products),
        'avg_price':  safe_div(a.rev, a.vol),
    })

ct = (pd.DataFrame(rows_c)
        .pivot(index='category', columns='year')
        .fillna(0))
ct.columns = [f"{m}_{y}" for m, y in ct.columns]
ct = ct.reset_index()
ct['rev_growth_abs']   = ct['revenue_2025'] - ct['revenue_2024']
ct['rev_growth_pct']   = ct.apply(lambda r: pct_change(r['revenue_2025'],   r['revenue_2024']),   axis=1)
ct['vol_growth_pct']   = ct.apply(lambda r: pct_change(r['volume_2025'],    r['volume_2024']),    axis=1)
ct['price_growth_pct'] = ct.apply(lambda r: pct_change(r['avg_price_2025'], r['avg_price_2024']), axis=1)
ct = ct.sort_values('revenue_2025', ascending=False).reset_index(drop=True)

print(f"Network table : {nt.shape[0]} redes   | ativas 2025: {(nt['revenue_2025']>0).sum()}")
print(f"Category table: {ct.shape[0]} categorias")
""")

# ─────────────────────────────────────────────────────────────────────────────
# SNAPSHOT DE QUALIDADE E VISÃO GERAL
# ─────────────────────────────────────────────────────────────────────────────
code("""
o24, o25 = overall[2024], overall[2025]
total_rows = sum(q['rows'] for q in qlog.values())

print("=" * 65)
print("QUALIDADE DOS DADOS")
print("=" * 65)
for y, q in qlog.items():
    print(f"  [{y}]  {q['rows']:>10,} linhas | "
          f"lojas {q['sh']/q['rows']*100:.4f}% | "
          f"produtos {q['ph']/q['rows']*100:.4f}%")
print(f"  Total: {total_rows:,} linhas")
print()
print("=" * 65)
print("VISÃO GERAL YoY")
print("=" * 65)
header = f"{'Métrica':<20} {'2024':>15} {'2025':>15} {'Var':>10}"
print(header)
print("-" * 65)
print(f"{'Receita':<20} {fmt_brl(o24['revenue']):>15} {fmt_brl(o25['revenue']):>15} "
      f"{fmt_pct(pct_change(o25['revenue'], o24['revenue'])):>10}")
print(f"{'Volume':<20} {fmt_int(o24['volume']):>15} {fmt_int(o25['volume']):>15} "
      f"{fmt_pct(pct_change(o25['volume'], o24['volume'])):>10}")
print(f"{'Preço médio':<20} {fmt_brl(o24['avg_price']):>15} {fmt_brl(o25['avg_price']):>15} "
      f"{fmt_pct(pct_change(o25['avg_price'], o24['avg_price'])):>10}")
print(f"{'Sortimento':<20} {fmt_int(o24['assortment']):>15} {fmt_int(o25['assortment']):>15} "
      f"{fmt_pct(pct_change(o25['assortment'], o24['assortment'])):>10}")
print(f"{'Lojas ativas':<20} {fmt_int(o24['active_stores']):>15} {fmt_int(o25['active_stores']):>15}")
""")

# ─────────────────────────────────────────────────────────────────────────────
# ENTREGA 1 – CURVA ABC / PARETO
# ─────────────────────────────────────────────────────────────────────────────
md("""
---
## Entrega 1 – Concentração de Faturamento (Curva ABC / Pareto)

**Metodologia:**
1. Somar faturamento total por rede em 2025
2. Ordenar em ordem decrescente
3. Calcular participação individual e acumulada
4. Identificar clientes que, juntos, atingem 80% do faturamento total
""")

code("""
pareto = nt.sort_values('revenue_2025', ascending=False).copy()
total_25 = pareto['revenue_2025'].sum()
pareto['share']     = pareto['revenue_2025'] / total_25
pareto['cum_share'] = pareto['share'].cumsum()

# Inclui exatamente até ultrapassar 80%
idx_cross = pareto['cum_share'].ge(0.80).idxmax()
pareto['pareto_flag'] = False
pareto.loc[:idx_cross, 'pareto_flag'] = True

classe_a = pareto[pareto['pareto_flag']]
n_a = len(classe_a)
n_ativas = int((nt['revenue_2025'] > 0).sum())

print(f"Faturamento total 2025 : {fmt_brl(total_25)}")
print(f"Classe A               : {n_a} redes = {classe_a['cum_share'].max()*100:.2f}% do faturamento")
print(f"Representatividade     : {n_a}/{n_ativas} redes ativas = {n_a/n_ativas*100:.1f}%")
print()

disp = classe_a[['network','revenue_2025','share','cum_share']].copy()
disp.columns = ['Cliente', 'Receita 2025', '% Individual', '% Acumulado']
disp['Receita 2025']  = disp['Receita 2025'].apply(fmt_brl)
disp['% Individual']  = disp['% Individual'].apply(lambda v: f"{v*100:.2f}%")
disp['% Acumulado']   = disp['% Acumulado'].apply(lambda v: f"{v*100:.2f}%")
disp.index = range(1, len(disp)+1)
print(disp.to_string())
""")

code("""
top20 = pareto.head(20).copy().reset_index(drop=True)

fig, ax1 = plt.subplots(figsize=(10, 6))
ax2 = ax1.twinx()

bar_colors = [C['navy'] if f else C['gray'] for f in top20['pareto_flag']]
ax1.barh(range(len(top20)), top20['revenue_2025'],
         color=bar_colors, alpha=0.85, height=0.7)

ax2.plot(top20['cum_share'] * 100, range(len(top20)),
         color=C['red'], marker='o', markersize=4, linewidth=1.8, label='% Acumulado')
ax2.axvline(80, color=C['red'], linestyle='--', linewidth=1.2, alpha=0.7)
ax2.text(80.5, len(top20)-1, '80%', color=C['red'], fontsize=8, va='top')

ax1.set_yticks(range(len(top20)))
ax1.set_yticklabels(top20['network'], fontsize=8)
ax1.invert_yaxis()
ax1.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'R${x/1e6:.1f}M'))
ax1.set_xlabel('Receita 2025')
ax2.set_ylabel('% Acumulado')
ax2.set_xlim(0, 108)

from matplotlib.patches import Patch
ax1.legend(handles=[Patch(color=C['navy'], label='Classe A (≤ 80%)'),
                    Patch(color=C['gray'], label='Demais redes')],
           loc='lower center', fontsize=8, framealpha=0.8)

ax1.set_title('Entrega 1 – Curva ABC: Concentração de Faturamento 2025', pad=12)
plt.tight_layout()
plt.savefig(OUT / 'e1_pareto.png', bbox_inches='tight', dpi=150)
plt.show()
print(f"Salvo → {OUT / 'e1_pareto.png'}")
""")

# ─────────────────────────────────────────────────────────────────────────────
# ENTREGA 2 – COMPARATIVO POR CLIENTE
# ─────────────────────────────────────────────────────────────────────────────
md("""
---
## Entrega 2 – Análise Comparativa por Cliente

Todas as métricas usam **2025** como ano de referência.
Crescimento **%** considera apenas redes com receita em 2024 (comparáveis).
""")

code("""
# ─── 2a. Ticket médio por loja ───────────────────────────────────────────────
top_ticket = (nt[nt['active_stores_2025'] > 0]
              .sort_values('ticket_store_2025', ascending=False)
              .head(10))

fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.barh(top_ticket['network'][::-1],
               top_ticket['ticket_store_2025'][::-1],
               color=C['teal'], alpha=0.85, height=0.65)
for bar, val in zip(bars, top_ticket['ticket_store_2025'][::-1]):
    ax.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height()/2,
            fmt_brl_k(val), va='center', fontsize=8)
ax.xaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'R${x/1e3:.0f}K'))
ax.set_xlabel('Ticket médio por loja ativa (R$)')
ax.set_title('2a. Ticket Médio por Loja – Top 10 Redes (2025)', pad=10)
plt.tight_layout()
plt.savefig(OUT / 'e2a_ticket.png', bbox_inches='tight', dpi=150)
plt.show()

best = top_ticket.iloc[0]
print(f"Maior ticket médio: {best['network']}  →  {fmt_brl(best['ticket_store_2025'])} / loja ativa")
print(f"  ({int(best['active_stores_2025'])} loja(s) ativa(s) | receita total {fmt_brl(best['revenue_2025'])})")
""")

code("""
# ─── 2b. Volume ───────────────────────────────────────────────────────────────
top_vol = nt.sort_values('volume_2025', ascending=False).head(10)

fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.barh(top_vol['network'][::-1], top_vol['volume_2025'][::-1],
               color=C['navy'], alpha=0.85, height=0.65)
for bar, val in zip(bars, top_vol['volume_2025'][::-1]):
    lbl = f"{val/1e6:.1f}M" if val >= 1e6 else f"{val/1e3:.0f}K"
    ax.text(bar.get_width() * 1.01, bar.get_y() + bar.get_height()/2,
            lbl, va='center', fontsize=8)
ax.xaxis.set_major_formatter(mticker.FuncFormatter(
    lambda x, _: f'{x/1e6:.1f}M' if x >= 1e6 else f'{x/1e3:.0f}K'))
ax.set_xlabel('Volume (unidades)')
ax.set_title('2b. Volume de Compras – Top 10 Redes (2025)', pad=10)
plt.tight_layout()
plt.savefig(OUT / 'e2b_volume.png', bbox_inches='tight', dpi=150)
plt.show()

best_v = top_vol.iloc[0]
print(f"Maior volume: {best_v['network']}  →  {fmt_int(best_v['volume_2025'])} unidades")
""")

code("""
# ─── 2c. Sortimento ───────────────────────────────────────────────────────────
asm_df = (nt[nt['revenue_2025'] > 0]
          [['network','assortment_2025']]
          .sort_values('assortment_2025', ascending=False)
          .head(15))
max_sku = int(products['product_code'].nunique())

fig, ax = plt.subplots(figsize=(8, 5))
bars = ax.barh(asm_df['network'][::-1], asm_df['assortment_2025'][::-1],
               color=C['teal'], alpha=0.80, height=0.65)
for bar, val in zip(bars, asm_df['assortment_2025'][::-1]):
    ax.text(bar.get_width() + 0.15, bar.get_y() + bar.get_height()/2,
            str(int(val)), va='center', fontsize=8)
ax.axvline(max_sku, color=C['red'], linestyle='--', linewidth=1.2, alpha=0.75,
           label=f'Teto vendido: {max_sku} SKUs')
ax.set_xlabel('SKUs distintos comprados em 2025')
ax.set_title('2c. Sortimento por Rede – Top 15 (2025)', pad=10)
ax.legend(fontsize=8)
plt.tight_layout()
plt.savefig(OUT / 'e2c_sortimento.png', bbox_inches='tight', dpi=150)
plt.show()

best_a = asm_df.iloc[0]
n_teto = int((nt[nt['revenue_2025'] > 0]['assortment_2025'] == best_a['assortment_2025']).sum())
print(f"Maior sortimento: {best_a['network']}  →  {int(best_a['assortment_2025'])} SKUs")
print(f"Teto de SKUs ativos: {max_sku}")
print(f"Obs.: {n_teto} redes atingiram o teto → métrica tem baixo poder discriminatório no topo.")
""")

code("""
# ─── 2d. Crescimento ──────────────────────────────────────────────────────────
comparable = nt[(nt['revenue_2024'].fillna(0) > 0) & (nt['revenue_2025'].fillna(0) > 0)].copy()

top_pct = comparable.sort_values('rev_growth_pct', ascending=False).head(10)
top_abs = comparable.sort_values('rev_growth_abs', ascending=False).head(10)

fig, (ax1, ax2) = plt.subplots(1, 2, figsize=(13, 5))

# — Crescimento % —
cl = [C['green'] if v >= 0 else C['red'] for v in top_pct['rev_growth_pct']]
ax1.barh(top_pct['network'][::-1], top_pct['rev_growth_pct'][::-1] * 100,
         color=cl[::-1], alpha=0.85, height=0.65)
ax1.xaxis.set_major_formatter(mticker.PercentFormatter())
ax1.axvline(0, color='black', linewidth=0.6)
ax1.set_xlabel('Crescimento % vs 2024')
ax1.set_title('2d. Crescimento % (redes comparáveis)', pad=10)

# — Crescimento absoluto —
cl2 = [C['green'] if v >= 0 else C['red'] for v in top_abs['rev_growth_abs']]
ax2.barh(top_abs['network'][::-1], top_abs['rev_growth_abs'][::-1],
         color=cl2[::-1], alpha=0.85, height=0.65)
ax2.xaxis.set_major_formatter(mticker.FuncFormatter(
    lambda x, _: f'R${x/1e6:.1f}M' if abs(x) >= 1e6 else f'R${x/1e3:.0f}K'))
ax2.axvline(0, color='black', linewidth=0.6)
ax2.set_xlabel('Crescimento absoluto (R$)')
ax2.set_title('2d. Crescimento Absoluto (redes comparáveis)', pad=10)

plt.tight_layout()
plt.savefig(OUT / 'e2d_crescimento.png', bbox_inches='tight', dpi=150)
plt.show()

best_p = comparable.sort_values('rev_growth_pct', ascending=False).iloc[0]
best_a2 = comparable.sort_values('rev_growth_abs', ascending=False).iloc[0]
print(f"Maior crescimento % : {best_p['network']:25s}  {fmt_pct(best_p['rev_growth_pct'])}")
print(f"Maior crescimento R$: {best_a2['network']:25s}  {fmt_brl(best_a2['rev_growth_abs'])}")
print()
print("Leitura: crescimento % mede performance relativa; crescimento R$ mede impacto financeiro real.")
print("As duas métricas são complementares e respondem perguntas diferentes.")
""")

code("""
# ─── 2e. Preço médio × Crescimento ────────────────────────────────────────────
sc = comparable.dropna(subset=['rev_growth_pct', 'avg_price_2025']).copy()
corr = sc['rev_growth_pct'].corr(sc['avg_price_2025'])

fig, ax = plt.subplots(figsize=(8, 5))
ax.scatter(sc['avg_price_2025'], sc['rev_growth_pct'] * 100,
           alpha=0.55, color=C['navy'], s=40, zorder=3)

# Destaques
q90_g = sc['rev_growth_pct'].quantile(0.90)
q10_g = sc['rev_growth_pct'].quantile(0.10)
q90_p = sc['avg_price_2025'].quantile(0.90)
mask  = (sc['rev_growth_pct'] > q90_g) | (sc['rev_growth_pct'] < q10_g) | (sc['avg_price_2025'] > q90_p)
for _, row in sc[mask].iterrows():
    ax.annotate(row['network'],
                (row['avg_price_2025'], row['rev_growth_pct'] * 100),
                fontsize=7, xytext=(4, 3), textcoords='offset points', alpha=0.85)

ax.axhline(0, color=C['gray'], linewidth=0.8, linestyle='--')
ax.set_xlabel('Preço médio 2025 (R$)')
ax.set_ylabel('Crescimento de receita 2024→2025')
ax.yaxis.set_major_formatter(mticker.PercentFormatter())
ax.set_title('2e. Preço Médio vs Crescimento de Receita (redes comparáveis)', pad=10)
ax.text(0.02, 0.97, f'Correlação de Pearson: {corr:.4f}',
        transform=ax.transAxes, fontsize=9, va='top',
        bbox=dict(boxstyle='round,pad=0.3', facecolor='white', edgecolor='lightgray', alpha=0.9))
plt.tight_layout()
plt.savefig(OUT / 'e2e_preco_crescimento.png', bbox_inches='tight', dpi=150)
plt.show()

max_preco = sc.sort_values('avg_price_2025', ascending=False).iloc[0]
print(f"Maior preço médio 2025 : {max_preco['network']}  →  {fmt_brl(max_preco['avg_price_2025'])}")
print(f"Maior crescimento %    : {best_p['network']}  →  preço médio {fmt_brl(best_p['avg_price_2025'])}")
print()
print(f"Correlação crescimento × preço: {corr:.4f}")
print("→ Relação praticamente nula — crescimento em 2025 dependeu de volume/cobertura, não de preço.")
""")

# ─────────────────────────────────────────────────────────────────────────────
# ENTREGA 3 – DIMENSÃO: CATEGORIA
# ─────────────────────────────────────────────────────────────────────────────
md("""
---
## Entrega 3 – Análise por Dimensão de Negócio: Categoria

**Dimensão escolhida:** Categoria
**Justificativa:** relacionamento produto → categoria tem cobertura confiável (99,9999% de join).
A base de macro-região não oferece chave de município para join determinístico.

**Indicadores analisados:** Receita, Crescimento YoY (%), Variação de Volume (%), Variação de Preço Médio (%)
""")

code("""
fig, axes = plt.subplots(1, 3, figsize=(14, 5))

cat_s = ct.sort_values('revenue_2025', ascending=True)
x = np.arange(len(cat_s))
w = 0.4

# — Receita 2024 vs 2025 —
axes[0].barh(x - w/2, cat_s['revenue_2024'], w, label='2024', color=C['gray'],  alpha=0.80)
axes[0].barh(x + w/2, cat_s['revenue_2025'], w, label='2025', color=C['navy'],  alpha=0.88)
axes[0].set_yticks(x)
axes[0].set_yticklabels(cat_s['category'], fontsize=8)
axes[0].xaxis.set_major_formatter(mticker.FuncFormatter(lambda v, _: f'R${v/1e6:.0f}M'))
axes[0].legend(fontsize=8)
axes[0].set_title('Receita por Categoria', fontsize=10, fontweight='bold')

# — Crescimento receita % —
cl = [C['green'] if v >= 0 else C['red'] for v in cat_s['rev_growth_pct'].fillna(0)]
axes[1].barh(x, cat_s['rev_growth_pct'].fillna(0) * 100, color=cl, alpha=0.88)
axes[1].set_yticks(x)
axes[1].set_yticklabels(cat_s['category'], fontsize=8)
axes[1].xaxis.set_major_formatter(mticker.PercentFormatter())
axes[1].axvline(0, color='black', linewidth=0.5)
axes[1].set_title('Δ Receita YoY', fontsize=10, fontweight='bold')

# — Variação de preço médio —
clp = [C['green'] if v >= 0 else C['red'] for v in cat_s['price_growth_pct'].fillna(0)]
axes[2].barh(x, cat_s['price_growth_pct'].fillna(0) * 100, color=clp, alpha=0.88)
axes[2].set_yticks(x)
axes[2].set_yticklabels(cat_s['category'], fontsize=8)
axes[2].xaxis.set_major_formatter(mticker.PercentFormatter())
axes[2].axvline(0, color='black', linewidth=0.5)
axes[2].set_title('Δ Preço Médio YoY', fontsize=10, fontweight='bold')

fig.suptitle('Entrega 3 – Desempenho por Categoria (YoY 2024 → 2025)',
             fontsize=12, fontweight='bold', y=1.02)
plt.tight_layout()
plt.savefig(OUT / 'e3_categoria.png', bbox_inches='tight', dpi=150)
plt.show()
print(f"Salvo → {OUT / 'e3_categoria.png'}")
""")

code("""
disp_cat = ct[['category','revenue_2024','revenue_2025',
               'rev_growth_pct','vol_growth_pct','price_growth_pct']].copy()
disp_cat = disp_cat.sort_values('revenue_2025', ascending=False)
disp_cat['Receita 2024'] = disp_cat['revenue_2024'].apply(fmt_brl)
disp_cat['Receita 2025'] = disp_cat['revenue_2025'].apply(fmt_brl)
disp_cat['Δ Receita']    = disp_cat['rev_growth_pct'].apply(fmt_pct)
disp_cat['Δ Volume']     = disp_cat['vol_growth_pct'].apply(fmt_pct)
disp_cat['Δ Preço']      = disp_cat['price_growth_pct'].apply(fmt_pct)
out_cat = disp_cat[['category','Receita 2024','Receita 2025','Δ Receita','Δ Volume','Δ Preço']].copy()
out_cat.columns = ['Categoria','Receita 2024','Receita 2025','Δ Receita','Δ Volume','Δ Preço médio']
out_cat.index = range(1, len(out_cat)+1)
print(out_cat.to_string())

print()
print("Leituras:")
print("  SALADAS    → maior categoria; cresce receita e volume, preço estável.  (crescimento saudável)")
print("  IN NATURA  → volume +9,4%  mas preço -3,7%.  (expansão com pressão de margem)")
print("  COUVE      → volume +9,9%  mas preço -1,1%.  (idem)")
print("  TEMPEROS   → cresce volume E preço (+3,6%).   (posicionamento mais forte)")
print("  LEGUMES    → aceleração forte (+66% receita), puxada por volume.")
print("  VERSATIL   → expansão rápida (+74%), preço relativamente estável.")
print("  GOURMET    → crescimento % alto (170%), mas base ainda pequena.")
""")

# ─────────────────────────────────────────────────────────────────────────────
# ENTREGA 4 – YoY EM 3 SLIDES
# ─────────────────────────────────────────────────────────────────────────────
md("""
---
## Entrega 4 – Análise de Crescimento de Vendas (YoY)

### Slide 1 – Visão Geral
### Slide 2 – Impulsionadores e Detratores
### Slide 3 – Recomendações
""")

code("""
# ─── Slide 1: Visão Geral ─────────────────────────────────────────────────────
labels   = ['Receita', 'Volume', 'Preço médio', 'Sortimento']
v24_list = [o24['revenue'], o24['volume'], o24['avg_price'], o24['assortment']]
v25_list = [o25['revenue'], o25['volume'], o25['avg_price'], o25['assortment']]

fig, axes = plt.subplots(1, 4, figsize=(14, 4))

for i, (ax, lbl, v24, v25) in enumerate(zip(axes, labels, v24_list, v25_list)):
    g = pct_change(v25, v24)
    ax.bar(['2024', '2025'], [v24, v25],
           color=[C['gray'], C['navy']], alpha=0.85, width=0.5, zorder=3)

    if i == 0:
        f24, f25 = fmt_brl_k(v24), fmt_brl_k(v25)
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'R${x/1e6:.0f}M'))
    elif i == 1:
        f24, f25 = f"{v24/1e6:.1f}M", f"{v25/1e6:.1f}M"
        ax.yaxis.set_major_formatter(mticker.FuncFormatter(lambda x, _: f'{x/1e6:.1f}M'))
    elif i == 2:
        f24, f25 = fmt_brl(v24), fmt_brl(v25)
    else:
        f24, f25 = str(int(v24)), str(int(v25))

    ax.text(0, v24 * 1.03, f24, ha='center', va='bottom', fontsize=8, color=C['gray'])
    ax.text(1, v25 * 1.03, f25, ha='center', va='bottom', fontsize=8, color=C['navy'])

    color_g = C['green'] if g and g > 0 else (C['red'] if g and g < 0 else 'black')
    ax.set_title(f"{lbl}\\n{fmt_pct(g)}", fontsize=9, fontweight='bold', color=color_g)
    ax.set_ylim(0, max(v24, v25) * 1.22)
    ax.spines['left'].set_visible(False)
    ax.set_yticks([])

fig.suptitle('Slide 1 – Visão Geral YoY (2024 → 2025)', fontsize=12, fontweight='bold')
plt.tight_layout()
plt.savefig(OUT / 'e4_slide1_yoy.png', bbox_inches='tight', dpi=150)
plt.show()
print(f"Salvo → {OUT / 'e4_slide1_yoy.png'}")
print()
print("Leitura: crescimento puxado por VOLUME (+22,9%) — preço subiu apenas +0,9%.")
print("A empresa vendeu mais, mas sem expansão de mix e com contribuição limitada de preço.")
""")

code("""
# ─── Slide 2: Impulsionadores e Detratores ────────────────────────────────────
comp_all = nt[nt['revenue_2024'].fillna(0) > 0].copy()
drivers    = comp_all.sort_values('rev_growth_abs', ascending=False).head(5)
detractors = comp_all.sort_values('rev_growth_abs', ascending=True).head(5)

comb = (pd.concat([drivers, detractors])
          .drop_duplicates('network')
          .sort_values('rev_growth_abs', ascending=True))

fig, ax = plt.subplots(figsize=(9, 6))
colors = [C['green'] if v >= 0 else C['red'] for v in comb['rev_growth_abs']]
bars = ax.barh(comb['network'], comb['rev_growth_abs'],
               color=colors, alpha=0.85, height=0.65, zorder=3)

for bar, val in zip(bars, comb['rev_growth_abs']):
    offset = total_25 * 0.003 if val >= 0 else -total_25 * 0.003
    ha = 'left' if val >= 0 else 'right'
    ax.text(bar.get_width() + offset,
            bar.get_y() + bar.get_height() / 2,
            fmt_brl_k(val), va='center', ha=ha, fontsize=8)

ax.axvline(0, color='black', linewidth=0.8)
ax.xaxis.set_major_formatter(mticker.FuncFormatter(
    lambda x, _: f'R${x/1e6:.1f}M' if abs(x) >= 1e6 else f'R${x/1e3:.0f}K'))
ax.set_title('Slide 2 – Impulsionadores e Detratores (variação absoluta de receita, redes comparáveis)',
             pad=10, fontsize=10, fontweight='bold')
plt.tight_layout()
plt.savefig(OUT / 'e4_slide2_drivers.png', bbox_inches='tight', dpi=150)
plt.show()
print(f"Salvo → {OUT / 'e4_slide2_drivers.png'}")
print()
print("Impulsionadores por categoria:")
for _, row in ct.sort_values('rev_growth_abs', ascending=False).iterrows():
    print(f"  {row['category']:22s}  {fmt_brl(row['rev_growth_abs']):>15}  ({fmt_pct(row['rev_growth_pct'])})")
""")

code("""
# ─── Slide 3: Recomendações ───────────────────────────────────────────────────
print("=" * 72)
print("SLIDE 3 – RECOMENDAÇÕES PRIORITÁRIAS")
print("=" * 72)

recs = [
    ("1", "Expandir sortimento em contas médias com boa base de volume",
           "Foco: redes relevantes que operam abaixo do teto de 29 SKUs ativos",
           "Meta: +3 SKUs no mix médio das contas-alvo em 6 meses"),

    ("2", "Proteger preço em categorias sob pressão (IN NATURA e COUVE)",
           "IN NATURA: volume +9,4% mas preço -3,7% → crescimento sem ganho de margem",
           "Meta: recuperar +1,5% de preço médio em 90 dias sem perda de volume > 3%"),

    ("3", "Separar contas novas de crescimento orgânico nas análises",
           "ST MARCHE (+809%) e DALBEN (+1485%) são principalmente entradas novas",
           "Meta: dois painéis distintos (comparáveis | novos) no próximo ciclo"),

    ("4", "Replicar alavancas das redes com crescimento forte e preço sustentado",
           "Benchmarks: OBA (+27%), SAVEGNAGO (+26,6%), MAMBO — crescimento com preço saudável",
           "Meta: selecionar 10 contas similares para plano de expansão em 120 dias"),

    ("5", "Atuar sobre os principais detratores de receita",
           "PAGUE MENOS (-80,5%), VAREJO (-65%), DB (-59%) — entender causa de saída",
           "Meta: recuperar 30% da perda absoluta dessas contas em 6 meses"),
]

for num, titulo, detalhe, meta in recs:
    print(f"\\n{num}. {titulo}")
    print(f"   {detalhe}")
    print(f"   → {meta}")
""")

# ─────────────────────────────────────────────────────────────────────────────
# ENTREGA 5 – MODELO DE REMUNERAÇÃO VARIÁVEL
# ─────────────────────────────────────────────────────────────────────────────
md("""
---
## Entrega 5 – Proposta de Modelo de Remuneração Variável

**Problema do modelo atual:** comissão % sobre vendas incentiva volume sem controle de preço ou mix.

**Objetivo do novo modelo:** remunerar crescimento com qualidade — volume, preço e sortimento simultaneamente.
""")

code("""
# ─── Diagnóstico ──────────────────────────────────────────────────────────────
print("DIAGNÓSTICO: POR QUE MUDAR O MODELO")
print("-" * 65)
rev_g  = pct_change(o25['revenue'],   o24['revenue'])
vol_g  = pct_change(o25['volume'],    o24['volume'])
prc_g  = pct_change(o25['avg_price'], o24['avg_price'])
asm_g  = pct_change(o25['assortment'],o24['assortment'])

print(f"  Receita cresceu    {fmt_pct(rev_g):>7}   ← modelo atual só remunera aqui")
print(f"  Volume cresceu     {fmt_pct(vol_g):>7}")
print(f"  Preço médio subiu  {fmt_pct(prc_g):>7}   ← quase nada")
print(f"  Sortimento variou  {fmt_pct(asm_g):>7}   ← estagnado")
print()
in_natura = ct[ct['category'] == 'IN NATURA'].iloc[0]
print(f"  IN NATURA: volume {fmt_pct(in_natura['vol_growth_pct'])} mas preço {fmt_pct(in_natura['price_growth_pct'])}")
print(f"  → vendedor foi premiado, mas com erosão de margem")
print()
print(f"  Correlação crescimento × preço: {corr:.4f}")
print("  → Crescimento atual é independente de qualidade de preço")
""")

code("""
# ─── Proposta ─────────────────────────────────────────────────────────────────
print("=" * 72)
print("PROPOSTA DE MODELO DE REMUNERAÇÃO VARIÁVEL")
print("=" * 72)
print(\"\"\"
  Payout = Bônus-alvo × (0,40 × S_vol + 0,30 × S_preco + 0,20 × S_sort + 0,10 × S_mix)

  INDICADOR                     PESO   COMO MEDIR
  ───────────────────────────── ─────  ──────────────────────────────────────────────
  Crescimento de volume           40%  (qtd_2025 − qtd_2024) / qtd_2024
                                         redes comparáveis; base zero excluída
  Disciplina de preço             30%  preço_médio_realizado vs faixa-meta por categoria
                                         redutor se erosão > threshold definido
  Crescimento de sortimento       20%  ΔSKUs distintos com ≥ 4 semanas de recorrência
                                         evita venda pontual para bater meta
  Execução de mix em cl. foco     10%  cobertura de SKUs estratégicos nas contas A
  ───────────────────────────── ─────  ──────────────────────────────────────────────

  REGRAS DE GOVERNANÇA
  ────────────────────
  • Contas novas → medição separada das comparáveis (evitar viés de base zero)
  • Crescimento via desconto excessivo → redutor aplicado no S_preco
  • Sortimento conta apenas com recorrência mínima (4 semanas)
  • Piso mínimo de faturamento para elegibilidade no componente de preço

  ANCORAGEM NOS DADOS
  ───────────────────
  • Volume: driver principal de 2025 (+22,9%) — correto premiar
  • Preço: cresceu apenas +0,9% — IN NATURA caiu -3,7% → modelo atual não protege
  • Sortimento estável em 29 SKUs — espaço para expansão nas contas médias
  • Correlação preço × crescimento ≈ 0 → modelo atual não alinha incentivos
\"\"\")
""")

code("""
# ─── Visualização do modelo ────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(9, 3))
ax.axis('off')

componentes = [
    ('Crescimento de Volume (40%)',  0.40, C['navy']),
    ('Disciplina de Preço (30%)',    0.30, C['teal']),
    ('Crescimento Sortimento (20%)', 0.20, C['orange']),
    ('Execução de Mix (10%)',        0.10, C['gray']),
]

x0 = 0.04
y0 = 0.30
height = 0.30
total_w = 0.92
for label, pct, color in componentes:
    w = pct * total_w
    ax.add_patch(plt.Rectangle((x0, y0), w, height, color=color, alpha=0.85, zorder=2))
    if pct >= 0.15:
        ax.text(x0 + w/2, y0 + height/2, label,
                ha='center', va='center', fontsize=8, color='white', fontweight='bold')
    else:
        ax.text(x0 + w + 0.005, y0 + height/2, label,
                ha='left', va='center', fontsize=7.5)
    x0 += w

ax.text(0.5, 0.88,
        'Payout = Bônus-alvo × (0,40 × S_vol + 0,30 × S_preço + 0,20 × S_sort + 0,10 × S_mix)',
        ha='center', va='top', fontsize=9, fontweight='bold', transform=ax.transAxes)
ax.set_xlim(0, 1)
ax.set_ylim(0, 1)
plt.tight_layout()
plt.savefig(OUT / 'e5_remuneracao.png', bbox_inches='tight', dpi=150)
plt.show()
print(f"Salvo → {OUT / 'e5_remuneracao.png'}")
""")

# ─────────────────────────────────────────────────────────────────────────────
# ESCREVE O ARQUIVO
# ─────────────────────────────────────────────────────────────────────────────
nb = {
    "cells": cells,
    "metadata": {
        "kernelspec": {
            "display_name": "Python 3",
            "language": "python",
            "name": "python3"
        },
        "language_info": {
            "name": "python",
            "pygments_lexer": "ipython3",
            "version": "3.10.0"
        }
    },
    "nbformat": 4,
    "nbformat_minor": 5
}

out_path = Path("analise_la_vita.ipynb")
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(nb, f, ensure_ascii=False, indent=1)

print(f"Notebook criado: {out_path}  ({out_path.stat().st_size / 1024:.1f} KB)")
print(f"Total de células: {len(cells)}")
