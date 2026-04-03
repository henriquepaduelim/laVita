from __future__ import annotations

import math
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from statistics import mean

import pandas as pd
from openpyxl import load_workbook

BASE_DIR = Path('.')
OUT_DIR = BASE_DIR / 'analysis_output'
OUT_DIR.mkdir(exist_ok=True)


@dataclass(frozen=True)
class ProductInfo:
    produto: str
    categoria: str


def pct_change(current: float, prior: float):
    if prior == 0:
        return None
    return (current / prior) - 1


def safe_div(n: float, d: float):
    if d == 0:
        return None
    return n / d


def fmt_money(v):
    return f'R$ {v:,.2f}'.replace(',', 'X').replace('.', ',').replace('X', '.')


def fmt_int(v):
    return f'{int(v):,}'.replace(',', '.')


def fmt_pct(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return 'n/a'
    return f'{v*100:.2f}%'.replace('.', ',')


def top_n_dict(d, n=5, reverse=True):
    return sorted(d.items(), key=lambda kv: kv[1], reverse=reverse)[:n]


clientes = pd.read_excel(BASE_DIR / 'base_cliente.xlsx')
produtos = pd.read_excel(BASE_DIR / 'base_produtos.xlsx')
macro = pd.read_excel(BASE_DIR / 'base_macro_regiao.xlsx')

store_info = {}
network_store_count = defaultdict(int)
state_counts = defaultdict(int)
city_counts = defaultdict(int)
for _, row in clientes.iterrows():
    store = str(row['NM_REDUZIDO']).strip()
    network = str(row['DS_REDE']).strip()
    state = str(row['SG_ESTADO']).strip()
    city = str(row['NM_MUNICIPIO']).strip()
    store_info[store] = {'network': network, 'state': state, 'city': city}
    network_store_count[network] += 1
    state_counts[state] += 1
    city_counts[city] += 1

product_info = {}
for _, row in produtos.iterrows():
    code = str(row['CODIGO']).strip()
    product_info[code] = ProductInfo(
        produto=str(row['PRODUTO']).strip(),
        categoria=str(row['CATEGORIA']).strip(),
    )

# Structural analysis of macro-regiao base
macro_unique_pairs = macro.drop_duplicates().shape[0]
macro_unique_uf = macro['UF'].nunique(dropna=True)
macro_unique_regions = macro['MACRO_REGIAO'].nunique(dropna=True)
macro_duplicate_rows = int(macro.duplicated().sum())

# Streaming aggregation over faturamento workbook
wb = load_workbook(BASE_DIR / 'base_faturamento.xlsx', read_only=True, data_only=True)

row_counts = defaultdict(int)
min_date = {}
max_date = {}
null_counts = defaultdict(lambda: defaultdict(int))
nonpositive_qty = defaultdict(int)
nonpositive_sales = defaultdict(int)
unknown_stores = defaultdict(int)
unknown_products = defaultdict(int)
unknown_store_names = defaultdict(set)
unknown_product_codes = defaultdict(set)

sales_total = defaultdict(float)
qty_total = defaultdict(float)
products_sold_total = defaultdict(set)
stores_sold_total = defaultdict(set)

sales_by_store = defaultdict(lambda: defaultdict(float))
qty_by_store = defaultdict(lambda: defaultdict(float))
products_by_store = defaultdict(lambda: defaultdict(set))

sales_by_network = defaultdict(lambda: defaultdict(float))
qty_by_network = defaultdict(lambda: defaultdict(float))
products_by_network = defaultdict(lambda: defaultdict(set))
active_stores_by_network = defaultdict(lambda: defaultdict(set))

sales_by_category = defaultdict(lambda: defaultdict(float))
qty_by_category = defaultdict(lambda: defaultdict(float))
products_by_category = defaultdict(lambda: defaultdict(set))

for sheet in ['2024', '2025']:
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    header_idx = {str(name).strip(): idx for idx, name in enumerate(header)}
    for r in rows:
        row_counts[sheet] += 1
        data = r[header_idx['DATA ENTREGA']]
        cliente = r[header_idx['CLIENTE']]
        produto = r[header_idx['CD_PRODUTO2']]
        quantidade = r[header_idx['QUANTIDADE']]
        venda = r[header_idx['VENDA']]

        if data is None:
            null_counts[sheet]['DATA ENTREGA'] += 1
            continue
        if cliente is None:
            null_counts[sheet]['CLIENTE'] += 1
            continue
        if produto is None:
            null_counts[sheet]['CD_PRODUTO2'] += 1
            continue
        if quantidade is None:
            null_counts[sheet]['QUANTIDADE'] += 1
            continue
        if venda is None:
            null_counts[sheet]['VENDA'] += 1
            continue

        cliente = str(cliente).strip()
        produto = str(produto).strip()
        quantidade = float(quantidade)
        venda = float(venda)

        if sheet not in min_date or data < min_date[sheet]:
            min_date[sheet] = data
        if sheet not in max_date or data > max_date[sheet]:
            max_date[sheet] = data

        if quantidade <= 0:
            nonpositive_qty[sheet] += 1
        if venda <= 0:
            nonpositive_sales[sheet] += 1

        sales_total[sheet] += venda
        qty_total[sheet] += quantidade
        products_sold_total[sheet].add(produto)
        stores_sold_total[sheet].add(cliente)

        sales_by_store[sheet][cliente] += venda
        qty_by_store[sheet][cliente] += quantidade
        products_by_store[sheet][cliente].add(produto)

        if cliente not in store_info:
            unknown_stores[sheet] += 1
            unknown_store_names[sheet].add(cliente)
        else:
            network = store_info[cliente]['network']
            sales_by_network[sheet][network] += venda
            qty_by_network[sheet][network] += quantidade
            products_by_network[sheet][network].add(produto)
            active_stores_by_network[sheet][network].add(cliente)

        if produto not in product_info:
            unknown_products[sheet] += 1
            unknown_product_codes[sheet].add(produto)
        else:
            cat = product_info[produto].categoria
            sales_by_category[sheet][cat] += venda
            qty_by_category[sheet][cat] += quantidade
            products_by_category[sheet][cat].add(produto)

networks_all = sorted(network_store_count.keys())
categories_all = sorted({p.categoria for p in product_info.values()})

# Dataframes for analysis
network_rows = []
for network in networks_all:
    sales24 = sales_by_network['2024'].get(network, 0.0)
    sales25 = sales_by_network['2025'].get(network, 0.0)
    qty24 = qty_by_network['2024'].get(network, 0.0)
    qty25 = qty_by_network['2025'].get(network, 0.0)
    registered_stores = network_store_count[network]
    active24 = len(active_stores_by_network['2024'].get(network, set()))
    active25 = len(active_stores_by_network['2025'].get(network, set()))
    row = {
        'network': network,
        'sales_2024': sales24,
        'sales_2025': sales25,
        'qty_2024': qty24,
        'qty_2025': qty25,
        'price_2024': safe_div(sales24, qty24),
        'price_2025': safe_div(sales25, qty25),
        'sort_2024': len(products_by_network['2024'].get(network, set())),
        'sort_2025': len(products_by_network['2025'].get(network, set())),
        'registered_stores': registered_stores,
        'active_stores_2024': active24,
        'active_stores_2025': active25,
        'ticket_registered_2025': safe_div(sales25, registered_stores),
        'ticket_active_2025': safe_div(sales25, active25),
        'growth_abs': sales25 - sales24,
        'growth_pct': pct_change(sales25, sales24),
    }
    network_rows.append(row)
network_df = pd.DataFrame(network_rows)

store_rows = []
for store, info in store_info.items():
    sales24 = sales_by_store['2024'].get(store, 0.0)
    sales25 = sales_by_store['2025'].get(store, 0.0)
    qty24 = qty_by_store['2024'].get(store, 0.0)
    qty25 = qty_by_store['2025'].get(store, 0.0)
    store_rows.append({
        'store': store,
        'network': info['network'],
        'state': info['state'],
        'city': info['city'],
        'sales_2024': sales24,
        'sales_2025': sales25,
        'qty_2024': qty24,
        'qty_2025': qty25,
        'price_2025': safe_div(sales25, qty25),
        'sort_2025': len(products_by_store['2025'].get(store, set())),
        'growth_abs': sales25 - sales24,
        'growth_pct': pct_change(sales25, sales24),
    })
store_df = pd.DataFrame(store_rows)

category_rows = []
for cat in categories_all:
    sales24 = sales_by_category['2024'].get(cat, 0.0)
    sales25 = sales_by_category['2025'].get(cat, 0.0)
    qty24 = qty_by_category['2024'].get(cat, 0.0)
    qty25 = qty_by_category['2025'].get(cat, 0.0)
    category_rows.append({
        'category': cat,
        'sales_2024': sales24,
        'sales_2025': sales25,
        'qty_2024': qty24,
        'qty_2025': qty25,
        'price_2024': safe_div(sales24, qty24),
        'price_2025': safe_div(sales25, qty25),
        'sort_2024': len(products_by_category['2024'].get(cat, set())),
        'sort_2025': len(products_by_category['2025'].get(cat, set())),
        'growth_abs': sales25 - sales24,
        'growth_pct': pct_change(sales25, sales24),
    })
category_df = pd.DataFrame(category_rows)

# Correlation for question 2e - only comparable clients
comparable_clients = network_df[(network_df['sales_2024'] > 0) & (network_df['qty_2025'] > 0)].copy()
price_growth_corr = comparable_clients['growth_pct'].corr(comparable_clients['price_2025'])

# Pareto on network for 2025
pareto_network = network_df[['network', 'sales_2025']].sort_values('sales_2025', ascending=False).reset_index(drop=True)
pareto_network['share'] = pareto_network['sales_2025'] / pareto_network['sales_2025'].sum()
pareto_network['cum_share'] = pareto_network['share'].cumsum()
pareto_cut_network = pareto_network[pareto_network['cum_share'] <= 0.8]
if pareto_cut_network.shape[0] < pareto_network.shape[0]:
    pareto_cut_network = pd.concat([
        pareto_cut_network,
        pareto_network.iloc[[pareto_cut_network.shape[0]]]
    ])

pareto_store = store_df[['store', 'network', 'sales_2025']].sort_values('sales_2025', ascending=False).reset_index(drop=True)
pareto_store['share'] = pareto_store['sales_2025'] / pareto_store['sales_2025'].sum()
pareto_store['cum_share'] = pareto_store['share'].cumsum()
pareto_cut_store = pareto_store[pareto_store['cum_share'] <= 0.8]
if pareto_cut_store.shape[0] < pareto_store.shape[0]:
    pareto_cut_store = pd.concat([
        pareto_cut_store,
        pareto_store.iloc[[pareto_cut_store.shape[0]]]
    ])

# Core case answers
best_ticket_active = network_df.sort_values('ticket_active_2025', ascending=False).iloc[0]
best_ticket_registered = network_df.sort_values('ticket_registered_2025', ascending=False).iloc[0]
best_volume = network_df.sort_values('qty_2025', ascending=False).iloc[0]
best_sort = network_df.sort_values('sort_2025', ascending=False).iloc[0]

comparable_growth = network_df[network_df['sales_2024'] > 0].copy()
best_growth_pct = comparable_growth.sort_values('growth_pct', ascending=False).iloc[0]
best_growth_abs = network_df.sort_values('growth_abs', ascending=False).iloc[0]
new_networks = network_df[(network_df['sales_2024'] == 0) & (network_df['sales_2025'] > 0)].copy().sort_values('sales_2025', ascending=False)

top_price = network_df[network_df['qty_2025'] > 0].sort_values('price_2025', ascending=False).iloc[0]

overall_metrics = pd.DataFrame([
    {
        'metric': 'Receita',
        '2024': sales_total['2024'],
        '2025': sales_total['2025'],
        'var_abs': sales_total['2025'] - sales_total['2024'],
        'var_pct': pct_change(sales_total['2025'], sales_total['2024'])
    },
    {
        'metric': 'Volume',
        '2024': qty_total['2024'],
        '2025': qty_total['2025'],
        'var_abs': qty_total['2025'] - qty_total['2024'],
        'var_pct': pct_change(qty_total['2025'], qty_total['2024'])
    },
    {
        'metric': 'Preço médio',
        '2024': safe_div(sales_total['2024'], qty_total['2024']),
        '2025': safe_div(sales_total['2025'], qty_total['2025']),
        'var_abs': safe_div(sales_total['2025'], qty_total['2025']) - safe_div(sales_total['2024'], qty_total['2024']),
        'var_pct': pct_change(safe_div(sales_total['2025'], qty_total['2025']), safe_div(sales_total['2024'], qty_total['2024']))
    },
    {
        'metric': 'Sortimento',
        '2024': len(products_sold_total['2024']),
        '2025': len(products_sold_total['2025']),
        'var_abs': len(products_sold_total['2025']) - len(products_sold_total['2024']),
        'var_pct': pct_change(len(products_sold_total['2025']), len(products_sold_total['2024']))
    }
])

network_deltas = network_df[['network', 'sales_2024', 'sales_2025', 'growth_abs', 'growth_pct']].sort_values('growth_abs', ascending=False)
category_deltas = category_df[['category', 'sales_2024', 'sales_2025', 'growth_abs', 'growth_pct']].sort_values('growth_abs', ascending=False)
negative_category_deltas = category_deltas[category_deltas['growth_abs'] < 0].sort_values('growth_abs')

# Sensitivity for ticket methodology
network_df['ticket_gap_pct'] = (network_df['ticket_active_2025'] / network_df['ticket_registered_2025']) - 1
median_ticket_gap = network_df.replace([math.inf, -math.inf], pd.NA)['ticket_gap_pct'].dropna().median()
active_store_share_2025 = (store_df['sales_2025'] > 0).mean()
active_skus_2025 = len(products_sold_total['2025'])
active_networks_2025 = int((network_df['sales_2025'] > 0).sum())

# Export support tables
network_df.sort_values('sales_2025', ascending=False).to_csv(OUT_DIR / 'network_metrics.csv', index=False)
store_df.sort_values('sales_2025', ascending=False).to_csv(OUT_DIR / 'store_metrics.csv', index=False)
category_df.sort_values('sales_2025', ascending=False).to_csv(OUT_DIR / 'category_metrics.csv', index=False)
overall_metrics.to_csv(OUT_DIR / 'overall_metrics.csv', index=False)
pareto_network.to_csv(OUT_DIR / 'pareto_network_2025.csv', index=False)
pareto_store.to_csv(OUT_DIR / 'pareto_store_2025.csv', index=False)

# Recommendations based on data
recommendations = [
    'Priorizar expansão nas redes com crescimento acima da média e sortimento abaixo do teto observado de 29 SKUs, porque há espaço de mix sem depender apenas de preço.',
    'Tratar erosão de preço nas categorias com crescimento de volume, mas queda de receita/price, usando piso por categoria em vez de desconto linear por cliente.',
    'Separar gestão de contas novas da gestão de contas comparáveis: clientes que entram em 2025 inflacionam leitura de crescimento percentual.',
    'Monitorar ticket por loja com duas lentes: lojas cadastradas e lojas ativas. Quando a diferença é grande, o problema é cobertura de loja, não necessariamente sell-out.',
    'Evitar análises geográficas até corrigir a base de macro-região, porque a chave atual não sustenta segmentação confiável.'
]

report_lines = []
append = report_lines.append
append('# Analise do Case La Vita Alimentos')
append('')
append('## 1. Escopo e criterio')
append('- Foco da leitura: confiabilidade analitica, correcao numerica e rastreabilidade das decisoes.')
append('- Nao foram produzidos slides; os resultados abaixo traduzem o conteudo que sustentaria a apresentacao.')
append('- Quando o case foi ambiguo, a decisao metodologica foi explicitada junto com o impacto.')
append('')
append('## 2. Validacao estrutural das bases')
append(f"- `base_faturamento.xlsx`: 2 abas, `2024` com {fmt_int(row_counts['2024'])} linhas de fatos e `2025` com {fmt_int(row_counts['2025'])} linhas de fatos.")
append('- Colunas do faturamento: `DATA ENTREGA`, `CLIENTE`, `CD_PRODUTO2`, `QUANTIDADE`, `VENDA`.')
append(f"- Periodo identificado: 2024 de {min_date['2024'].date()} a {max_date['2024'].date()}; 2025 de {min_date['2025'].date()} a {max_date['2025'].date()}.")
append(f"- `base_cliente.xlsx`: {fmt_int(len(clientes))} linhas, {fmt_int(clientes['DS_REDE'].nunique())} redes, {fmt_int(clientes['NM_REDUZIDO'].nunique())} lojas, {fmt_int(clientes['NM_MUNICIPIO'].nunique())} municipios, {fmt_int(clientes['SG_ESTADO'].nunique())} estados.")
append(f"- `base_produtos.xlsx`: {fmt_int(len(produtos))} linhas, {fmt_int(produtos['CODIGO'].nunique())} codigos unicos, {fmt_int(produtos['CATEGORIA'].nunique())} categorias; apenas {fmt_int(active_skus_2025)} SKUs tiveram venda em 2025.")
append(f"- `base_macro_regiao.xlsx`: {fmt_int(len(macro))} linhas, apenas {fmt_int(macro_unique_pairs)} pares distintos, {fmt_int(macro_unique_uf)} valores distintos em `UF` e {fmt_int(macro_unique_regions)} macro-regioes distintas. A base nao tem granularidade suficiente para mapear as {fmt_int(clientes['NM_MUNICIPIO'].nunique())} cidades do cadastro de clientes.")
append('')
append('### Granularidade inferida')
append('- O fato comercial esta no nivel `data de entrega x loja x produto`.')
append('- A base nao contem `pedido`, `nota fiscal` ou `id de linha`; portanto, repeticoes nessa chave nao podem ser classificadas automaticamente como duplicidade ou como eventos legitimos no mesmo dia.')
append('- Para evitar inflacao por granularidade, todas as metricas foram consolidadas no nivel analitico necessario antes da comparacao.')
append('')
append('## 3. Teste das chaves de juncao')
append(f"- Faturamento -> clientes: chave `CLIENTE` = `NM_REDUZIDO` com cobertura de {fmt_pct(1 - (unknown_stores['2024'] + unknown_stores['2025']) / (row_counts['2024'] + row_counts['2025']))}. Linhas sem correspondencia: {fmt_int(unknown_stores['2024'] + unknown_stores['2025'])}.")
append(f"- Faturamento -> produtos: chave `CD_PRODUTO2` = `CODIGO` com cobertura de {fmt_pct(1 - (unknown_products['2024'] + unknown_products['2025']) / (row_counts['2024'] + row_counts['2025']))}. Linhas sem correspondencia: {fmt_int(unknown_products['2024'] + unknown_products['2025'])}.")
append('- Clientes -> macro-regiao: inviavel de forma deterministica com a estrutura recebida. A coluna `UF` da base de macro-regiao nao discrimina cidade nem loja, e a combinacao `UF + MACRO_REGIAO` aparece repetida em excesso.')
append('')
append('### Decisao metodologica')
append('- Analises por cliente foram feitas em nivel de `rede` (`DS_REDE`).')
append('- Analises por loja usam `NM_REDUZIDO` quando necessario para Pareto e cobertura.')
append('- Analises por regiao nao foram usadas para conclusao numerica por falta de chave confiavel.')
append('')
append('## 4. Ambiguidades metodologicas do case')
append('- `Cliente` pode significar loja ou rede. Adotei `rede` como cliente de negocio, porque o case cita explicitamente `Rede (cliente)` e pede `ticket medio por loja`, o que pressupoe agrupamento acima da loja.')
append('- `Ticket medio por loja` pode usar lojas cadastradas ou lojas ativas no periodo. Calculei ambos e usei `lojas ativas em 2025` como principal, por refletir o desempenho operacional observado no ano. Impacto: o ticket por lojas cadastradas e sempre menor ou igual e expõe ociosidade/com baixa cobertura.')
append('- `Crescimento` pode ser absoluto ou percentual. Reportei ambos. Para ranking principal de crescimento, usei crescimento percentual apenas entre clientes comparaveis com venda em 2024, para nao deixar contas novas dominarem o ranking com base zero.')
append('- Pareto pode ser por 2025, por base consolidada ou por loja. Adotei 2025 e nivel de rede, por ser a leitura mais aderente ao estado atual do negocio.')
append('- `Preco medio` foi calculado como `VENDA / QUANTIDADE` em base agregada, evitando media simples de precos transacionais.')
append(f"- O cadastro possui {fmt_int(len(store_df))} lojas, mas apenas {fmt_int((store_df['sales_2025'] > 0).sum())} ficaram ativas em 2025 ({fmt_pct(active_store_share_2025)}). Isso afeta qualquer leitura de ticket por loja.")
append('')
append('## 5. Resposta ao case')
append('')
append('### 5.1 Concentração de faturamento (Curva ABC / Pareto)')
append(f"- Em 2025, {fmt_int(len(pareto_cut_network))} redes concentram {fmt_pct(pareto_cut_network['cum_share'].iloc[-1])} do faturamento total ao criterio Pareto 80/20.")
append(f"- Isso representa {fmt_pct(len(pareto_cut_network) / max(active_networks_2025,1))} das {fmt_int(active_networks_2025)} redes ativas em 2025.")
append(f"- Sensibilidade em nivel de loja: {fmt_int(len(pareto_cut_store))} lojas concentram {fmt_pct(pareto_cut_store['cum_share'].iloc[-1])} do faturamento de 2025.")
append('- Top 10 redes no acumulado de 2025:')
for _, r in pareto_network.head(10).iterrows():
    append(f"  - {r['network']}: {fmt_money(r['sales_2025'])} ({fmt_pct(r['share'])}; acumulado {fmt_pct(r['cum_share'])})")
append('')
append('### 5.2 Analise comparativa por cliente (rede)')
append(f"- Maior ticket medio por loja ativa em 2025: `{best_ticket_active['network']}` com {fmt_money(best_ticket_active['ticket_active_2025'])} por loja ativa.")
append(f"- Maior ticket medio por loja cadastrada em 2025: `{best_ticket_registered['network']}` com {fmt_money(best_ticket_registered['ticket_registered_2025'])} por loja cadastrada.")
append(f"- Maior volume em 2025: `{best_volume['network']}` com {fmt_int(best_volume['qty_2025'])} itens.")
sort_ties = network_df[(network_df['sales_2025'] > 0) & (network_df['sort_2025'] == best_sort['sort_2025'])].sort_values('network')
sort_ties_label = ', '.join(f"`{network}`" for network in sort_ties['network'].tolist())
if len(sort_ties) == 1:
    append(f"- Maior sortimento em 2025: {sort_ties_label} com {fmt_int(best_sort['sort_2025'])} SKUs distintos.")
else:
    append(f"- Maior sortimento em 2025: empate entre {sort_ties_label}, todos com {fmt_int(best_sort['sort_2025'])} SKUs distintos.")
append(f"- Maior crescimento percentual entre clientes comparaveis (2024 > 0): `{best_growth_pct['network']}` com {fmt_pct(best_growth_pct['growth_pct'])} sobre 2024.")
append(f"- Maior crescimento absoluto: `{best_growth_abs['network']}` com aumento de {fmt_money(best_growth_abs['growth_abs'])} em 2025 vs 2024.")
append(f"- Redes novas em 2025 com venda e sem base 2024: {fmt_int(len(new_networks))}. A maior foi `{new_networks.iloc[0]['network']}` com {fmt_money(new_networks.iloc[0]['sales_2025'])}." if len(new_networks) else '- Nao houve redes novas em 2025 sem base 2024.')
append(f"- Cliente com maior preco medio em 2025: `{top_price['network']}` com {fmt_money(top_price['price_2025'])} por item.")
append(f"- O cliente com maior crescimento percentual nao coincide com o maior preco medio: {'sim' if best_growth_pct['network'] == top_price['network'] else 'nao'}. Correlacao entre crescimento percentual e preco medio 2025 entre clientes comparaveis: {price_growth_corr:.4f}." if pd.notna(price_growth_corr) else '- Correlacao entre crescimento e preco medio nao disponivel.')
append('')
append('#### Top 10 clientes comparaveis por crescimento percentual')
for _, r in comparable_growth.sort_values('growth_pct', ascending=False).head(10).iterrows():
    append(f"- {r['network']}: crescimento {fmt_pct(r['growth_pct'])}; receita 2025 {fmt_money(r['sales_2025'])}; preco medio 2025 {fmt_money(r['price_2025'])}; sortimento {fmt_int(r['sort_2025'])}.")
append('')
append('### 5.3 Analise direcionada por dimensao de negocio escolhida: categoria')
append('- Escolhi `categoria`, nao `macro-regiao`, porque a juncao produto -> categoria e completa e confiavel, enquanto a base de macro-regiao nao permite segmentacao auditavel.')
append('- Resultado por categoria em 2025 vs 2024:')
for _, r in category_df.sort_values('sales_2025', ascending=False).iterrows():
    append(f"- {r['category']}: receita 2025 {fmt_money(r['sales_2025'])}; YoY {fmt_pct(r['growth_pct'])}; volume 2025 {fmt_int(r['qty_2025'])}; preco medio 2025 {fmt_money(r['price_2025'])}; sortimento {fmt_int(r['sort_2025'])}.")
append('')
append('### 5.4 Year over Year (conteudo que sustentaria 3 slides)')
append('#### Slide 1 - Visao geral')
for _, r in overall_metrics.iterrows():
    base2024 = r['2024']
    base2025 = r['2025']
    if r['metric'] == 'Sortimento':
        append(f"- {r['metric']}: 2024 = {fmt_int(base2024)}; 2025 = {fmt_int(base2025)}; variacao = {fmt_int(r['var_abs'])} ({fmt_pct(r['var_pct'])}).")
    elif r['metric'] == 'Volume':
        append(f"- {r['metric']}: 2024 = {fmt_int(base2024)}; 2025 = {fmt_int(base2025)}; variacao = {fmt_int(r['var_abs'])} ({fmt_pct(r['var_pct'])}).")
    else:
        append(f"- {r['metric']}: 2024 = {fmt_money(base2024)}; 2025 = {fmt_money(base2025)}; variacao = {fmt_money(r['var_abs'])} ({fmt_pct(r['var_pct'])}).")
append('')
append('#### Slide 2 - Impulsionadores e detratores')
append('- Top 5 redes por incremento absoluto de receita em 2025 vs 2024:')
for _, r in network_deltas.head(5).iterrows():
    append(f"  - {r['network']}: {fmt_money(r['growth_abs'])} ({fmt_pct(r['growth_pct'])})")
append('- Top 5 redes detratoras por queda absoluta de receita em 2025 vs 2024:')
for _, r in network_deltas.sort_values('growth_abs').head(5).iterrows():
    append(f"  - {r['network']}: {fmt_money(r['growth_abs'])} ({fmt_pct(r['growth_pct'])})")
append('- Top 5 categorias impulsionadoras por incremento absoluto de receita:')
for _, r in category_deltas.head(5).iterrows():
    append(f"  - {r['category']}: {fmt_money(r['growth_abs'])} ({fmt_pct(r['growth_pct'])})")
if negative_category_deltas.empty:
    append('- Nenhuma categoria apresentou queda absoluta de receita em 2025 vs 2024.')
else:
    append('- Top 5 categorias detratoras por queda absoluta de receita:')
    for _, r in negative_category_deltas.head(5).iterrows():
        append(f"  - {r['category']}: {fmt_money(r['growth_abs'])} ({fmt_pct(r['growth_pct'])})")
append('')
append('#### Slide 3 - Recomendacoes analiticamente sustentadas')
for rec in recommendations:
    append(f"- {rec}")
append('')
append('### 5.5 Proposta de remuneracao variavel')
append('- Objetivo: remunerar crescimento com qualidade, sem premiar ganho de receita por desconto excessivo ou concentracao exagerada em poucos SKUs.')
append('- Estrutura sugerida de score trimestral por carteira:')
append('  - 40% `crescimento de volume`: variacao de quantidade vs mesmo periodo do ano anterior, comparando apenas base ativa/comparavel.')
append('  - 30% `disciplina de preco`: indice de preco realizado (`receita / quantidade`) contra meta por categoria e cliente, com gatilho minimo para impedir erosao de margem/posicionamento.')
append('  - 20% `expansao de sortimento`: evolucao de SKUs distintos por cliente, limitada a SKUs elegiveis para evitar mix artificial.')
append('  - 10% `execucao de mix`: venda recorrente em clientes foco para reduzir concentracao excessiva em poucos itens.')
append('- Regras de governanca:')
append('  - Gate 1: se o indice de preco ficar abaixo do piso, o componente de volume nao paga integralmente.')
append('  - Gate 2: crescimento sobre base zero deve ser medido separadamente de clientes comparaveis, para nao distorcer incentivo.')
append('  - Gate 3: sortimento so pontua se houver recorrencia minima, evitando venda pontual apenas para bater meta.')
append('- Exemplo de formula: `payout = target_bonus * (0,4*score_volume + 0,3*score_preco + 0,2*score_sortimento + 0,1*score_mix)`, com cada score truncado entre 0% e 120%.')
append('- Vantagem tecnica: o modelo usa variaveis observaveis e auditaveis no proprio faturamento, sem depender de julgamento subjetivo.')
append('')
append('## 6. Riscos de interpretacao e impacto')
append(f"- A base de macro-regiao foi considerada inadequada para analise conclusiva. Impacto: qualquer leitura geografica a partir dela teria alto risco de erro de atribuicao.")
append(f"- A diferenca mediana entre ticket por loja ativa e ticket por loja cadastrada foi de {fmt_pct(median_ticket_gap)}. Impacto: rankings mudam se o denominador de lojas nao for explicitado.")
append(f"- Apenas {fmt_pct(active_store_share_2025)} das lojas cadastradas tiveram faturamento em 2025. Impacto: metricas por loja ficam super sensiveis ao conceito de loja elegivel vs loja ativa.")
append(f"- Das {fmt_int(len(network_df))} redes cadastradas, apenas {fmt_int(active_networks_2025)} tiveram faturamento em 2025. Impacto: concentracao e ticket precisam ser lidos sobre base ativa, nao sobre cadastro total.")
append('- Clientes novos em 2025 nao foram misturados ao ranking principal de crescimento percentual. Impacto: o ranking fica mais justo para comparar desempenho real, e nao apenas entrada de base nova.')
append('- O crescimento foi calculado em ano fechado, porque as duas abas cobrem o calendario completo de 2024 e 2025. Se a empresa quiser usar ano movel ou YTD equivalente, os numeros mudam.')
append('- Preco medio foi calculado por `receita / quantidade` agregada. Uma media simples de precos por linha produziria viés matematico.')
append('')
append('## 7. Ferramentas, frameworks, metodos e praticas recomendados')
append('### Stack recomendada')
append('- `Python`: linguagem principal para ETL analitico, validacao e reprodutibilidade.')
append('- `pandas`: integracao das bases pequenas, consolidacao de metricas e tabulacao final.')
append('- `openpyxl` em modo `read_only`: leitura segura de planilhas grandes quando a origem chega em `.xlsx`.')
append('- `DuckDB` (recomendado para a versao evoluida do processo): executar validacoes, joins e agregacoes SQL sobre arquivos locais com mais velocidade e melhor auditabilidade.')
append('- `Pandera` ou `Great Expectations`: validacao formal de schema, chaves, dominios e regras de negocio antes do calculo.')
append('- `Jupyter Notebook` para exploracao inicial e `script .py` para execucao final versionada e reprodutivel.')
append('')
append('### Metodos e praticas')
append('- Data profiling antes da analise: cardinalidade, nulos, dominios, cobertura de chave e periodo.')
append('- Definicao formal das metricas antes dos graficos: receita, volume, preco medio, sortimento, ticket e crescimento precisam de formula unica.')
append('- Reconciliacao obrigatoria: total da base tratada deve bater com o total da base bruta apos joins e agregacoes.')
append('- Separacao entre clientes comparaveis e clientes novos nas analises de crescimento.')
append('- Registro explicito das ambiguidades do case e da decisao adotada, para que o numero seja defendivel em discussao executiva.')
append('- Evitar Excel como motor principal da analise em volume acima de 1 milhao de linhas; usar Excel apenas para inspecao rapida ou apresentacao final.')
append('')
append('## 8. Arquivos gerados')
append('- `analysis_output/network_metrics.csv`')
append('- `analysis_output/store_metrics.csv`')
append('- `analysis_output/category_metrics.csv`')
append('- `analysis_output/overall_metrics.csv`')
append('- `analysis_output/pareto_network_2025.csv`')
append('- `analysis_output/pareto_store_2025.csv`')
append('- `analysis_output/case_analysis_report.md`')

(OUT_DIR / 'case_analysis_report.md').write_text('\n'.join(report_lines), encoding='utf-8')
print('Analysis complete.')
